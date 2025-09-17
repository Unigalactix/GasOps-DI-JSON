#!/usr/bin/env python3
"""
Object-Oriented PDF Document Intelligence Processor
Converts PDF files to JSON using Azure Document Intelligence and AI processing.
Enhanced with standalone MTR CLI functionality for API-based processing.

Usage: python pdf_processor_oop.py

Architecture:
- PDFProcessor: Main orchestrator class
- DocumentIntelligenceOCR: OCR extraction class  
- AITemplateProcessor: AI processing class for JSON generation
- APIProcessor: Enhanced API-based PDF retrieval with embedded functionality
- XLSXProcessor: Excel file management with color-coded data merging

Features:
- Local PDF file processing
- API-based PDF retrieval by HeatNumber
- Embedded authentication and certificate handling
- Automatic JSON to XLSX conversion with color logic
- Enhanced error handling and logging
"""

import os
import sys
import json
import re
import time
import requests
import base64
import tempfile
import io
from dotenv import load_dotenv
from typing import Optional, Dict, Any
from datetime import datetime
from pathlib import Path

load_dotenv()

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import certificate API client
try:
    from scripts.cert_api_client import api_client
    CERT_API_AVAILABLE = True
except ImportError:
    CERT_API_AVAILABLE = False

# Add requests_pkcs12 for enhanced API calling
try:
    import requests_pkcs12
    REQUESTS_PKCS12_AVAILABLE = True
except ImportError:
    REQUESTS_PKCS12_AVAILABLE = False

# Add Azure Document Intelligence imports
try:
    from azure.core.credentials import AzureKeyCredential
    from azure.ai.documentintelligence import DocumentIntelligenceClient
    from azure.ai.documentintelligence.models import AnalyzeResult
    AZURE_DI_AVAILABLE = True
except ImportError:
    AZURE_DI_AVAILABLE = False

# Embedded API calling functionality
def call_weld_api(api_name, parameters, auth_token, pfx_source="./certificate/oamsapicert2023.pfx"):
    """Embedded API client for external welding management system endpoints."""

    api_endpoints = {
        "GetMTRFileDatabyHeatNumber": "/api/AIMTRMetaData/GetMTRFileDatabyHeatNumber",
    }

    endpoint = api_endpoints.get(api_name)
    if not endpoint:
        return {"error": f"Unknown API: {api_name}"}

    url = f"https://oamsapi.gasopsiq.com{endpoint}"
    payload = parameters

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "auth-token": auth_token
    }

    temp_file = None
    try:
        # If the input is a base64 string (not a file path), decode and save as temp file
        if not os.path.isfile(pfx_source):
            try:
                cert_bytes = base64.b64decode(pfx_source)
            except Exception as decode_err:
                return {"error": f"Failed to decode base64 certificate: {decode_err}"}
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pfx")
            temp_file.write(cert_bytes)
            temp_file.close()
            pfx_path = temp_file.name
        else:
            pfx_path = pfx_source

        with open(pfx_path, "rb") as f:
            pfx_data = f.read()

        # GET request for MTR API
        response = requests_pkcs12.get(
            url,
            headers=headers,
            params=payload,
            pkcs12_data=pfx_data,
            pkcs12_password="password1234"
        )

        try:
            result = response.json()
            return {"success": True, "data": result, "status_code": response.status_code}
        except Exception:
            return {"success": True, "data": response.text, "status_code": response.status_code}

    except Exception as e:
        return {"error": str(e)}
    finally:
        if temp_file:
            try:
                os.unlink(temp_file.name)
            except Exception:
                pass

# Embedded MTR tools functionality
def GetMTRFileDatabyHeatNumber(heat_number=None, company_mtr_file_id=None, auth_token=None):
    """Embedded MTR data retrieval function."""
    parameters = {
        "heatNumber": heat_number,
        "companyMTRFileID": company_mtr_file_id
    }
    # Clean parameters by removing None values
    parameters = {k: v for k, v in parameters.items() if v is not None}
    return call_weld_api("GetMTRFileDatabyHeatNumber", parameters, auth_token)

# Default authentication token
DEFAULT_AUTH_TOKEN = "OS8xNy8yMDI1IDc6MzM6MzYgUE0mNDgwJkNFREVNT05FVzAzMTQmOS8xNi8yMDI1IDc6MzM6MzYgUE0mQ0VERU1P"


class DocumentIntelligenceOCR:
    """Handles OCR text extraction using Azure Document Intelligence."""
    
    def __init__(self, endpoint: str, api_key: str, model_id: str = "prebuilt-document", api_version: str = "2023-07-31"):
        """Initialize OCR processor with Azure credentials."""
        self.endpoint = endpoint.rstrip('/')
        self.api_key = api_key
        self.model_id = model_id
        self.api_version = api_version
        
        if not self.endpoint or not self.api_key:
            raise ValueError("Missing Azure Document Intelligence credentials")
    
    def extract_text_from_pdf(self, file_bytes: bytes, content_type: str = "application/pdf") -> str:
        """Extract text from PDF using Document Intelligence OCR."""
        print(f"Starting OCR extraction with model: {self.model_id}")
        
        # Call Document Intelligence API
        ocr_result = self._call_document_intelligence_api(file_bytes, content_type)
        
        # Extract text from result
        extracted_text = self._parse_ocr_result(ocr_result)
        
        if not extracted_text.strip():
            raise RuntimeError("No text could be extracted from the document")
        
        print(f"Successfully extracted {len(extracted_text)} characters of text")
        return extracted_text
    
    def _call_document_intelligence_api(self, file_bytes: bytes, content_type: str) -> Dict[str, Any]:
        """Make API call to Document Intelligence service."""
        analyze_url = f"{self.endpoint}/formrecognizer/documentModels/{self.model_id}:analyze?api-version={self.api_version}"
        headers = {
            "Ocp-Apim-Subscription-Key": self.api_key,
            "Content-Type": content_type
        }
        
        # Submit analysis request
        resp = requests.post(analyze_url, headers=headers, data=file_bytes)
        
        if resp.status_code not in (200, 202):
            self._handle_api_error(resp)
        
        # Check if we have operation location for polling
        op_location = resp.headers.get("operation-location") or resp.headers.get("Operation-Location")
        if not op_location:
            return resp.json()
        
        # Poll for completion
        return self._poll_for_completion(op_location)
    
    def _handle_api_error(self, response: requests.Response):
        """Handle API error responses with helpful hints."""
        if response.status_code == 403:
            try:
                body = response.json()
                svc_msg = body.get("error", {}).get("message", response.text)
            except Exception:
                svc_msg = response.text
            
            hint = (
                "Access denied (403). This commonly means your Document Intelligence resource has "
                "Virtual Network or Firewall restrictions. Remedies:\n"
                "1) In Azure Portal: Document Intelligence resource → Networking → Enable public access\n"
                "2) Add your client IP to the allowed IP list\n"
                "3) Configure Private Endpoint with proper DNS if using VNet"
            )
            raise RuntimeError(f"OCR API call failed: {response.status_code} {svc_msg}\n{hint}")
        
        raise RuntimeError(f"OCR API call failed: {response.status_code} {response.text}")
    
    def _poll_for_completion(self, operation_location: str, max_retries: int = 60) -> Dict[str, Any]:
        """Poll the operation location until analysis is complete."""
        print("Waiting for OCR analysis to complete...")
        
        for attempt in range(max_retries):
            time.sleep(1)
            
            get_resp = requests.get(
                operation_location, 
                headers={"Ocp-Apim-Subscription-Key": self.api_key}
            )
            
            if get_resp.status_code not in (200, 201):
                raise RuntimeError(f"Polling failed: {get_resp.status_code} {get_resp.text}")
            
            result = get_resp.json()
            status = result.get("status", "").lower()
            
            if status == "succeeded":
                print("OCR analysis completed successfully")
                return result
            elif status in ("failed", "cancelled"):
                raise RuntimeError(f"OCR analysis {status}: {result}")
        
        raise RuntimeError("Timed out waiting for OCR analysis to complete")
    
    def _parse_ocr_result(self, result_json: Dict[str, Any]) -> str:
        """Extract plain text from Document Intelligence OCR result."""
        text_parts = []
        
        def recurse_text(obj):
            if isinstance(obj, dict):
                # Check for text content in various fields
                for key in ("content", "text", "value"):
                    if key in obj and isinstance(obj[key], str):
                        text_parts.append(obj[key])
                # Recurse through other keys
                for v in obj.values():
                    recurse_text(v)
            elif isinstance(obj, list):
                for item in obj:
                    recurse_text(item)
        
        recurse_text(result_json)
        return "\n".join(text_parts)


class AITemplateProcessor:
    """Handles AI processing to convert OCR text into structured JSON."""
    
    def __init__(self):
        """Initialize AI processor with available credentials."""
        self.ai_config = self._detect_ai_configuration()
        if not self.ai_config:
            raise ValueError("No AI configuration found. Please configure Azure OpenAI or OpenAI credentials.")
    
    def _detect_ai_configuration(self) -> Optional[Dict[str, str]]:
        """Detect and validate available AI configuration."""
        # Check Azure OpenAI first
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        azure_key = os.getenv("AZURE_OPENAI_KEY") or os.getenv("AZURE_OPENAI_API_KEY")
        azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")
        
        if azure_endpoint and azure_key and azure_deployment:
            return {
                "type": "azure_openai",
                "endpoint": azure_endpoint.rstrip('/'),
                "key": azure_key,
                "deployment": azure_deployment,
                "api_version": os.getenv("AZURE_OPENAI_API_VERSION", "2023-10-01")
            }
        
        # Check OpenAI
        openai_key = os.getenv("OPENAI_API_KEY")
        if openai_key:
            return {
                "type": "openai",
                "key": openai_key,
                "model": "gpt-3.5-turbo"
            }
        
        return None
    
    def load_template(self, template_path: Optional[str] = None) -> Dict[str, Any]:
        """Load and clean the JSON template."""
        if not template_path:
            # Try default locations
            template_paths = [
                r"C:\Users\kodag\Downloads\GITHUB\GasOps-DI-JSON\Sample json\sample.json",
                os.path.join(os.path.dirname(__file__), "Sample json", "sample.json")
            ]
        else:
            template_paths = [template_path]
        
        for path in template_paths:
            try:
                if os.path.exists(path):
                    with open(path, 'r', encoding='utf-8') as f:
                        template = json.load(f)
                    print(f"Loaded template from: {path}")
                    return self._clean_template_values(template)
            except Exception as e:
                print(f"Warning: Could not load template from {path}: {e}")
                continue
        
        print("Warning: Could not load template from any location, using fallback")
        return self._get_fallback_template()
    
    def _clean_template_values(self, obj: Any) -> Any:
        """Recursively clean template values, replacing sample data with null/empty values."""
        if isinstance(obj, dict):
            cleaned = {}
            for key, value in obj.items():
                if isinstance(value, (dict, list)):
                    cleaned[key] = self._clean_template_values(value)
                elif isinstance(value, str):
                    cleaned[key] = ""
                elif isinstance(value, (int, float)):
                    cleaned[key] = None
                elif isinstance(value, bool):
                    cleaned[key] = None
                else:
                    cleaned[key] = None
            return cleaned
        elif isinstance(obj, list):
            if len(obj) > 0:
                return [self._clean_template_values(obj[0])]
            else:
                return []
        else:
            return obj
    
    def _get_fallback_template(self) -> Dict[str, Any]:
        """Return a minimal fallback template structure."""
        return {
            "CompanyMTRFileID": None,
            "HeatNumber": "",
            "ZNumber": "",
            "CertificationDate": "",
            "HNPipeDetails": [{
                "PipeNumber": "",
                "Grade": "",
                "HNPipeHeatChemicalResults": {},
                "HNPipeChemicalCompResults": {},
                "HNPipeTensileTestResults": {},
                "HNPipeCVNResults": {},
                "HNPipeHardnessResults": {}
            }]
        }
    
    def process_text_to_json(self, extracted_text: str, template: Dict[str, Any], timeout: int = 30) -> Optional[Dict[str, Any]]:
        """Process extracted text into structured JSON using AI."""
        print("Processing text with AI to generate structured JSON...")
        
        system_msg = self._build_system_message()
        user_msg = self._build_user_message(template, extracted_text)
        
        try:
            response_content = self._call_ai_api(system_msg, user_msg, timeout)
            if not response_content:
                return None
            
            # Parse JSON from AI response
            parsed_json = self._extract_json_from_response(response_content)
            
            if parsed_json:
                print("Successfully generated structured JSON")
                return parsed_json
            else:
                print("Warning: Could not parse valid JSON from AI response")
                return None
                
        except Exception as e:
            print(f"AI processing failed: {e}")
            return None
    
    def _build_system_message(self) -> str:
        """Build the system message for AI processing."""
        return (
            "You are an expert AI assistant designed to extract and structure data from OCR-extracted content of Material Test Reports (MTRs). "
            "You will receive text content from an MTR document and a JSON schema. Your task is to accurately parse the provided text and populate the JSON schema with the corresponding values.\n\n"
            "Instructions:\n\n"
            "Identify Key Information: Analyze the input text to find specific data points like HeatNumber, CertificationDate, PipeManufacturerName, and various chemical and mechanical test results (e.g., YieldStrength, UltimateTensileStrength, HeatC, Product1Mn).\n\n"
            "Match and Map: Map the extracted values to the corresponding keys in the provided JSON schema.\n\n"
            "Handle Missing Data: If a specific data point is not found in the text, leave its corresponding value in the JSON as null. Do not create new keys or alter the structure.\n\n"
            "Preserve Structure: Maintain the exact JSON structure, including nested objects and arrays.\n\n"
            "Format Values:\n\n"
            "For numerical values, ensure they are represented as strings.\n\n"
            "For units, include them exactly as found in the source document.\n\n"
            "For dates, convert them to the MM/DD/YYYY format if necessary.\n\n"
            "Validate: Cross-reference the extracted values with the provided criteria (e.g., HeatCEPcmCriteria) to ensure they meet the specifications. Note any discrepancies.\n\n"
            "Chemical Equivalency (VERY IMPORTANT):\n"
            "When populating HNPipeChemicalEquivResults fields (for example: Product1CEPcm, ProductCEPcmCriteria, Product1CEIIW, ProductCEIIWCriteria, Product2CEPcm, Product2CEIIW), follow these strict rules:\n"
            "1) Match by explicit labels: Prefer values that are directly labeled in the document as 'CE (Pcm)', 'C.E. (Pcm)', 'CE Pcm', 'CEIIW', 'CE (IIW)', 'CE IIW', 'Carbon Equivalent (IIW)' or similar. If the document provides 'Product 1' / 'Product 2' or 'Pipe 1' / 'Pipe 2' labels, map the CE values to the corresponding product index (Product1 -> Product1 fields, Product2 -> Product2 fields).\n"
            "2) Nearest-label rule: If a CE value is not explicitly tied to 'Product1' or 'Product2', select the CE value that appears nearest in the text to the product/pipe identifier, table row, or the chemical composition block for that product.\n"
            "3) Do NOT swap products: Avoid assigning Product2's CE to Product1 and vice versa. If there are multiple CE values and you cannot confidently map them to a product by label or proximity, leave the ambiguous fields null.\n"
            "4) Normalization and format: Return numeric CE values as strings, preserving decimals. If the source shows a leading decimal like '.354', normalize to '0.354' for consistency. Do not add units.\n"
            "5) Criteria fields: For fields named '*Criteria' (e.g., ProductCEPcmCriteria), populate them only when the document explicitly shows a specification or criteria value. If no explicit criteria is present, leave the criteria field null.\n"
            "6) Verification: After extracting CE values, echo a short clarifying note in the JSON under an additional top-level key 'ExtractionNotes' (optional) only when you had to pick between ambiguous matches — otherwise omit this key. The primary output must remain the template structure.\n\n"
            "Tensile Results (IMPORTANT):\n"
            "When extracting tensile/mechanical test results (for example: YieldStrength, UltimateTensileStrength, YTRatio, SeamWeldTensileStrength and their unit fields), follow these strict rules:\n"
            "1) Prefer explicit labels: Look for labels such as 'Yield Strength', 'YS', 'YieldStrength', 'Yield (ksi)', 'Ultimate Tensile Strength', 'UTS', 'Seam Weld Tensile', 'Seam Weld', or similar. Map them to the corresponding fields exactly.\n"
            "2) Units: Capture units separately when they are provided (for example 'ksi'). Populate the '*Unit' field exactly as shown (e.g., 'ksi'). If a unit is missing but other values use 'ksi', infer 'ksi' only when confident; otherwise leave the unit null.\n"
            "3) YT Ratio: If a 'Y/T' or 'YTRatio' is provided (yield divided by tensile), capture the numeric string. Normalize leading decimals (e.g., '.77' -> '0.77').\n"
            "4) Numeric format: Return numeric values as strings, preserving one or two decimal places as found in the source. If the source uses a leading decimal, normalize to a leading zero ('.77' -> '0.77'). Do not append units to numeric fields — units must go into the separate '*Unit' fields.\n"
            "5) Table and proximity mapping: If results appear in a table or grouped block, map values by row/column association. Use proximity to associate a unit with its numeric value when the unit is shown once for the row.\n"
            "6) Seam weld values: Look for explicit 'Seam' or 'Seam Weld' qualifiers and map them to 'SeamWeldTensileStrength' and 'SeamWeldTensileStrengthUnits'. If only a single tensile value is present and seam weld isn't called out, do not invent seam weld entries — leave them null.\n"
            "7) Ambiguity: If more than one candidate value exists and you cannot confidently map them (by label, row, or proximity), leave the ambiguous fields null rather than guessing. Optionally include a short 'ExtractionNotes' key at top-level explaining the ambiguity.\n\n"
            "Your output must be a single, valid JSON object that adheres to the provided schema with the values replaced by the extracted data."
        )
    
    def _build_user_message(self, template: Dict[str, Any], text: str) -> str:
        """Build the user message with template and OCR text."""
        return (
            f"JSON TEMPLATE:\n{json.dumps(template, indent=2)}\n\n"
            f"OCR TEXT:\n{text[:50000]}\n\n"
            "INSTRUCTIONS (READ CAREFULLY):\n"
            "1) Output: Return ONLY a single, valid JSON object that matches the provided template structure. Do NOT output any additional text, explanation, or commentary.\n"
            "2) Use source data only: Replace template values only with data explicitly found in the OCR text. Do not invent values or use placeholder/sample values from the template.\n"
            "3) Do not change keys or structure: Preserve the exact keys, nesting, and array structure from the template. If a field cannot be found, set it to null (or empty string where appropriate per the template).\n"
            "4) Numeric formatting: All numeric fields must be returned as strings. Normalize leading decimals by adding a leading zero (e.g., '.354' -> '0.354', '.77' -> '0.77'). Preserve the number of decimals as shown in the source when possible.\n"
            "5) Units: When a unit (e.g., 'ksi') is provided in the source, populate the corresponding '*Unit' field exactly as shown. Do NOT append units to numeric fields. If a unit is not present, leave the '*Unit' field null.\n"
            "6) Dates: Convert any dates to MM/DD/YYYY format. If a date cannot be parsed confidently, leave the date field null.\n"
            "7) Chemical Equivalency mapping: For 'HNPipeChemicalEquivResults' fields (Product1CEPcm, ProductCEPcmCriteria, Product1CEIIW, ProductCEIIWCriteria, Product2CEPcm, Product2CEIIW): prefer explicitly labeled CE values (e.g., 'CE (Pcm)', 'CE Pcm', 'CEIIW', 'CE IIW'). Map values by explicit product labels first. If labels are not explicit, map by proximity to the product's chemical composition block or table row. Do NOT swap Product1 and Product2 values. If mapping is ambiguous, set the field to null.\n"
            "8) Criteria fields: Only populate '*Criteria' fields when the document explicitly lists a criteria/specification value. Otherwise leave them null.\n"
            "9) Tensile/mechanical mapping: For YieldStrength, UltimateTensileStrength, YTRatio, SeamWeldTensileStrength and their unit fields: prefer explicit labels (e.g., 'Yield Strength', 'YS', 'Ultimate Tensile Strength', 'UTS', 'Seam Weld Tensile'). Capture units separately into '*Unit' fields. Normalize numeric formats (leading zeros) and return numeric values as strings. Do not create seam-weld values if the document does not call them out.\n"
            "10) Tables and proximity: If data is in a table, respect row/column associations. Use nearest-label and table-row association to map values when explicit inline labels are missing.\n"
            "11) Ambiguity policy: If you cannot confidently map a value according to the rules above, set the field to null. Optionally include a short top-level key 'ExtractionNotes' with concise reasons when you purposely left fields null due to ambiguity (keep this note minimal).\n"
            "12) JSON validity: Ensure the returned JSON is syntactically valid (no trailing commas, correct quoting). Numeric strings should remain quoted.\n"
            "13) Final check: Before returning, ensure the object exactly matches the template keys and nesting; do not add extra metadata except the optional 'ExtractionNotes' when necessary.\n\n"
            "Return the populated JSON object now."
        )
    
    def _call_ai_api(self, system_msg: str, user_msg: str, timeout: int) -> Optional[str]:
        """Make API call to the configured AI service."""
        payload = {
            "messages": [
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
            "temperature": 0,
            "max_tokens": 4000,
        }
        
        if self.ai_config["type"] == "azure_openai":
            return self._call_azure_openai(payload, timeout)
        elif self.ai_config["type"] == "openai":
            return self._call_openai(payload, timeout)
        else:
            raise ValueError(f"Unknown AI configuration type: {self.ai_config['type']}")
    
    def _call_azure_openai(self, payload: Dict[str, Any], timeout: int) -> Optional[str]:
        """Call Azure OpenAI API."""
        url = (f"{self.ai_config['endpoint']}/openai/deployments/{self.ai_config['deployment']}/"
               f"chat/completions?api-version={self.ai_config['api_version']}")
        headers = {"api-key": self.ai_config["key"], "Content-Type": "application/json"}
        
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
        if resp.status_code not in (200, 201):
            raise RuntimeError(f"Azure OpenAI API call failed: {resp.status_code} {resp.text}")
        
        body = resp.json()
        return body.get("choices", [])[0].get("message", {}).get("content")
    
    def _call_openai(self, payload: Dict[str, Any], timeout: int) -> Optional[str]:
        """Call OpenAI API."""
        url = "https://api.openai.com/v1/chat/completions"
        headers = {"Authorization": f"Bearer {self.ai_config['key']}", "Content-Type": "application/json"}
        payload["model"] = self.ai_config["model"]
        
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
        if resp.status_code not in (200, 201):
            raise RuntimeError(f"OpenAI API call failed: {resp.status_code} {resp.text}")
        
        body = resp.json()
        return body.get("choices", [])[0].get("message", {}).get("content")
    
    def _extract_json_from_response(self, response: str) -> Optional[Dict[str, Any]]:
        """Extract and parse JSON from AI response."""
        # Try to find JSON object first
        starts = [m.start() for m in re.finditer(r"\{", response)]
        for start in starts:
            depth = 0
            for i in range(start, len(response)):
                if response[i] == "{":
                    depth += 1
                elif response[i] == "}":
                    depth -= 1
                    if depth == 0:
                        candidate = response[start:i+1]
                        try:
                            return json.loads(candidate)
                        except Exception:
                            break
        
        # If no object found, try array
        starts = [m.start() for m in re.finditer(r"\[", response)]
        for start in starts:
            depth = 0
            for i in range(start, len(response)):
                if response[i] == "[":
                    depth += 1
                elif response[i] == "]":
                    depth -= 1
                    if depth == 0:
                        candidate = response[start:i+1]
                        try:
                            return json.loads(candidate)
                        except Exception:
                            break
        
        # Fallback: try to parse the entire response
        try:
            return json.loads(response)
        except Exception:
            return None


class APIProcessor:
    """Handles API-based PDF retrieval and processing with enhanced functionality."""
    
    def __init__(self):
        """Initialize API processor with certificate configuration."""
        if not REQUESTS_PKCS12_AVAILABLE:
            raise ImportError("requests_pkcs12 is required for API processing. Install with: pip install requests-pkcs12")
        
        # API configuration
        self.api_base_url = "https://oamsapi.gasopsiq.com"
        self.api_endpoint = "/api/AIMTRMetaData/GetMTRFileDatabyHeatNumber"
        self.pfx_source = "./certificate/oamsapicert2023.pfx"
        self.pfx_password = "password1234"
        self.default_auth_token = DEFAULT_AUTH_TOKEN
        
        # Check if certificate exists
        if not os.path.exists(self.pfx_source):
            raise FileNotFoundError(f"Certificate file not found: {self.pfx_source}")
    
    def fetch_pdf_by_heat_number(self, heat_number: str, auth_token: str = None) -> bytes:
        """
        Fetch PDF data from API using HeatNumber with enhanced API calling.
        Uses the embedded call_weld_api and GetMTRFileDatabyHeatNumber functions.
        
        Args:
            heat_number: The heat number to search for
            auth_token: Authentication token (uses default if None)
            
        Returns:
            PDF file content as bytes
        """
        # Use default token if none provided
        if not auth_token:
            auth_token = self.default_auth_token
        
        print(f"Fetching MTR data for HeatNumber: {heat_number}")
        
        try:
            # Use embedded API function
            tool_result = GetMTRFileDatabyHeatNumber(
                heat_number=heat_number,
                company_mtr_file_id=None,
                auth_token=auth_token
            )
            
            if not tool_result.get("success"):
                raise RuntimeError(f"API call failed: {tool_result.get('error', 'Unknown error')}")
            
            # Process MTR document response
            if isinstance(tool_result.get("data"), dict):
                data = tool_result["data"]
                if "Obj" in data and data["Obj"]:
                    first_obj = data["Obj"][0]
                    binary_string = first_obj.get("BinaryString")
                    
                    if binary_string:
                        # Convert binary string to PDF bytes
                        try:
                            pdf_data = base64.b64decode(binary_string)
                            print(f"Successfully decoded binary string as base64 ({len(pdf_data)} bytes)")
                            return pdf_data
                        except Exception:
                            # Try as raw data if base64 fails
                            pdf_data = binary_string.encode('latin-1') if isinstance(binary_string, str) else binary_string
                            print(f"Using binary string as raw data ({len(pdf_data)} bytes)")
                            return pdf_data
                    else:
                        raise RuntimeError("No binary string found in API response")
                else:
                    raise RuntimeError("No MTR data found in API response")
            else:
                raise RuntimeError("Invalid API response format")
                
        except Exception as e:
            raise RuntimeError(f"Failed to fetch PDF from API: {e}")
    
    def convert_binary_to_pdf(self, binary_string: str, heat_number: str, save_locally: bool = True) -> str:
        """Convert binary string to PDF file with local storage option"""
        try:
            # Same logic as standalone script
            try:
                pdf_data = base64.b64decode(binary_string)
                print(f"Successfully decoded binary string as base64 for heat number {heat_number}")
            except Exception:
                pdf_data = binary_string.encode('latin-1') if isinstance(binary_string, str) else binary_string
                print(f"Using binary string as raw data for heat number {heat_number}")

            if save_locally:
                # Save to Sample json folder
                sample_json_dir = os.path.join(os.path.dirname(__file__), "Sample json")
                os.makedirs(sample_json_dir, exist_ok=True)
                pdf_path = os.path.join(sample_json_dir, f"{heat_number}.pdf")
            else:
                # Temp directory
                temp_dir = tempfile.gettempdir()
                pdf_path = os.path.join(temp_dir, f"{heat_number}.pdf")

            with open(pdf_path, 'wb') as pdf_file:
                pdf_file.write(pdf_data)

            print(f"PDF file created successfully: {pdf_path}")
            return pdf_path

        except Exception as e:
            print(f"Failed to convert binary to PDF for heat number {heat_number}: {str(e)}")
            raise Exception(f"PDF conversion failed: {str(e)}")
    
    def process_heat_number_to_json(self, heat_number: str, output_dir: Optional[str] = None, auth_token: str = None) -> tuple:
        """
        Fetch PDF by heat number, process it, and return paths.
        
        Args:
            heat_number: The heat number to process
            output_dir: Optional output directory for JSON file
            auth_token: Authentication token (uses default if None)
            
        Returns:
            Tuple of (temp_pdf_path, output_json_path)
        """
        # Fetch PDF data from API using enhanced method
        pdf_content = self.fetch_pdf_by_heat_number(heat_number, auth_token)
        
        # Create temporary PDF file
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
            temp_pdf.write(pdf_content)
            temp_pdf_path = temp_pdf.name
        
        # Set output path
        if output_dir:
            output_path = os.path.join(output_dir, f"{heat_number}.json")
        else:
            # Save to Sample json folder by default
            sample_json_dir = os.path.join(os.path.dirname(__file__), "Sample json")
            os.makedirs(sample_json_dir, exist_ok=True)
            output_path = os.path.join(sample_json_dir, f"{heat_number}.json")
        
        return temp_pdf_path, output_path


class XLSXProcessor:
    """Handles XLSX file creation and data merging with color logic."""
    
    def __init__(self, xlsx_template_path: str):
        """Initialize XLSX processor with template path."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX processing. Install with: pip install openpyxl")
        
        self.xlsx_template_path = Path(xlsx_template_path)
        
    def flatten_json(self, obj: Any, parent_key: str = "", sep: str = ".") -> Dict[str, str]:
        """
        Flatten a JSON-like object into a single-level dict with dot-separated keys.
        Creates smart mapping for Excel template headers.
        """
        items = {}
        if isinstance(obj, dict):
            for k, v in obj.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, (dict, list)):
                    items.update(self.flatten_json(v, new_key, sep=sep))
                else:
                    # Add both hierarchical and direct key
                    items[new_key] = str(v) if v is not None else ""
                    # Also add just the field name for direct Excel header matching
                    items[k] = str(v) if v is not None else ""
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                new_key = f"{parent_key}{sep}{i}" if parent_key else str(i)
                if isinstance(v, (dict, list)):
                    items.update(self.flatten_json(v, new_key, sep=sep))
                else:
                    items[new_key] = str(v) if v is not None else ""
        else:
            items[parent_key] = str(obj) if obj is not None else ""
        return items
    
    def update_xlsx_from_json(self, json_path: str) -> str:
        """
        Update XLSX file with JSON data according to merging rules.
        
        Args:
            json_path: Path to the JSON file to process
            
        Returns:
            Path to the updated XLSX file
        """
        json_path = Path(json_path)
        
        # Load JSON data
        with open(json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)
        
        # Create XLSX path in same directory as JSON with NEW_.xlsx name
        xlsx_path = json_path.parent / "NEW_.xlsx"
        
        # Check if XLSX file already exists
        if xlsx_path.exists():
            print(f"XLSX file already exists, updating: {xlsx_path}")
            # Load existing workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            template_ws = None  # No template needed for existing file
            
        elif self.xlsx_template_path.exists():
            print(f"Creating new XLSX file from template: {xlsx_path}")
            # Create new workbook from template
            template_wb = load_workbook(self.xlsx_template_path)
            template_ws = template_wb.active
            
            wb = Workbook()
            ws = wb.active
            
            # Copy first two rows from template with formatting
            self._copy_header_rows(template_ws, ws)
            
        else:
            raise FileNotFoundError(f"Template XLSX file not found: {self.xlsx_template_path}")
        
        # Clean up empty rows between headers and data before adding new data
        self._remove_empty_rows_after_headers(ws)
        
        # Get headers from second row (field names, not group headers)
        header_row = [cell.value for cell in ws[2]]
        
        # Flatten JSON data
        flat_data = self.flatten_json(json_data)
        row_to_process = [flat_data.get(str(h), "") for h in header_row]
        
        # Find HeatNumber column index
        heat_col_idx = self._find_heat_number_column(header_row)
        
        # Search for existing row with matching HeatNumber
        match_row = self._find_matching_row(ws, heat_col_idx, row_to_process)
        
        # Determine target row (overwrite existing or append new)
        target_row = match_row if match_row else ws.max_row + 1
        
        # Update row with color logic
        self._update_row_with_colors(ws, template_ws if template_ws else ws, 
                                   target_row, header_row, row_to_process)
        
        # Clean up any empty rows that might have been created during processing
        self._remove_empty_rows_after_headers(ws)
        
        # Save workbook
        wb.save(xlsx_path)
        
        action = "Updated existing" if xlsx_path.exists() else "Created new"
        print(f"{action} XLSX file: {xlsx_path}")
        return str(xlsx_path)
    
    def _copy_header_rows(self, source_ws, target_ws):
        """Copy first two rows from source to target with formatting."""
        for row_idx in [1, 2]:
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row_idx, column=col_idx)
                target_cell = target_ws.cell(row=row_idx, column=col_idx, value=source_cell.value)
                
                # Skip style copying to avoid openpyxl style errors
                # Just copy values for now
        
        # Freeze top two rows
        target_ws.freeze_panes = "A3"
    
    def _remove_empty_rows_after_headers(self, ws):
        """Remove empty rows between headers and data, moving data up."""
        print("Checking for empty rows between headers and data...")
        
        # Start checking from row 3 (after headers)
        rows_to_delete = []
        
        # Find all empty rows starting from row 3
        current_row = 3
        while current_row <= ws.max_row:
            # Check if the entire row is empty
            is_empty_row = True
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=current_row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    is_empty_row = False
                    break
            
            if is_empty_row:
                rows_to_delete.append(current_row)
                print(f"Found empty row: {current_row}")
            else:
                # If we hit a non-empty row, stop looking for consecutive empty rows
                # But continue if there might be more empty rows later
                pass
                
            current_row += 1
        
        # Delete empty rows (in reverse order to maintain row indices)
        for row_num in reversed(rows_to_delete):
            print(f"Deleting empty row: {row_num}")
            ws.delete_rows(row_num)
        
        if rows_to_delete:
            print(f"Removed {len(rows_to_delete)} empty rows, data moved up")
        else:
            print("No empty rows found between headers and data")
    
    def _find_heat_number_column(self, header_row) -> Optional[int]:
        """Find the column index for HeatNumber field."""
        for idx, col in enumerate(header_row):
            if str(col).lower() in ["heatnumber", "heat number", "heat_number"]:
                return idx
        return None
    
    def _find_matching_row(self, ws, heat_col_idx: Optional[int], row_data) -> Optional[int]:
        """Find existing row with matching HeatNumber."""
        if heat_col_idx is None:
            return None
        
        target_heat_number = row_data[heat_col_idx]
        if not target_heat_number:
            return None
        
        for row_idx in range(3, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=heat_col_idx + 1).value
            if str(cell_value) == str(target_heat_number):
                return row_idx
        
        return None
    
    def _update_row_with_colors(self, ws, template_ws, target_row: int, header_row, row_data):
        """Update row with data and apply color logic."""
        for col_idx, (header, value) in enumerate(zip(header_row, row_data), 1):
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            
            # Get previous value for color comparison
            prev_value = None
            has_color = False
            
            # For new files or when template is available, use template for comparison
            if template_ws and target_row <= template_ws.max_row:
                template_cell = template_ws.cell(row=target_row, column=col_idx)
                prev_value = template_cell.value
                # Check if template cell has any color
                if template_cell.fill and template_cell.fill.patternType and template_cell.fill.patternType != 'none':
                    has_color = True
            
            # Apply color logic - simplified
            if str(value) == str(prev_value) and has_color:
                # Value matches and template had color - copy the fill directly
                try:
                    cell.fill = template_cell.fill
                except Exception:
                    # If copying fails, set to default
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                # Value different or no previous color - set to white
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


class PDFProcessor:
    """Main orchestrator class for PDF document processing."""
    
    def __init__(self):
        """Initialize the PDF processor with OCR and AI components."""
        # Load configuration
        self.config = self._load_configuration()
        
        # Initialize components
        self.ocr_processor = DocumentIntelligenceOCR(
            endpoint=self.config["azure_di_endpoint"],
            api_key=self.config["azure_di_key"],
            model_id=self.config.get("azure_di_model_id", "prebuilt-document"),
            api_version=self.config.get("azure_di_api_version", "2023-07-31")
        )
        
        self.ai_processor = AITemplateProcessor()
        
        # Initialize API processor
        if REQUESTS_PKCS12_AVAILABLE:
            try:
                self.api_processor = APIProcessor()
                print("API processor initialized successfully")
            except Exception as e:
                self.api_processor = None
                print(f"Warning: API processor initialization failed: {e}")
        else:
            self.api_processor = None
            print("Warning: requests_pkcs12 not available. API processing disabled.")
        
        # Initialize XLSX processor
        xlsx_template_path = os.path.join(os.path.dirname(__file__), "Sample json", "pdf_test_output (TEST).xlsx")
        if OPENPYXL_AVAILABLE and os.path.exists(xlsx_template_path):
            self.xlsx_processor = XLSXProcessor(xlsx_template_path)
        else:
            self.xlsx_processor = None
            if not OPENPYXL_AVAILABLE:
                print("Warning: openpyxl not available. XLSX processing disabled.")
            else:
                print(f"Warning: XLSX template not found at {xlsx_template_path}. XLSX processing disabled.")
        
        print("PDF Processor initialized successfully")
    
    def _load_configuration(self) -> Dict[str, str]:
        """Load and validate configuration from environment variables."""
        config = {}
        
        # Azure Document Intelligence configuration
        config["azure_di_endpoint"] = (
            os.getenv("AZURE_DI_ENDPOINT") or 
            os.getenv("AZURE_FORM_RECOGNIZER_ENDPOINT")
        )
        config["azure_di_key"] = (
            os.getenv("AZURE_DI_KEY") or 
            os.getenv("AZURE_FORM_RECOGNIZER_KEY")
        )
        config["azure_di_model_id"] = os.getenv("AZURE_DI_MODEL_ID", "prebuilt-document")
        config["azure_di_api_version"] = os.getenv("AZURE_DI_API_VERSION", "2023-07-31")
        
        # Validate required configuration
        if not config["azure_di_endpoint"] or not config["azure_di_key"]:
            raise ValueError(
                "Missing Azure Document Intelligence credentials. "
                "Please configure AZURE_DI_ENDPOINT and AZURE_DI_KEY in .env file."
            )
        
        return config
    
    def process_pdf(self, pdf_path: str, output_path: Optional[str] = None, template_path: Optional[str] = None) -> str:
        """
        Process a PDF file and return the output JSON file path.
        
        Args:
            pdf_path: Path to the input PDF file
            output_path: Optional custom output path for JSON file
            template_path: Optional custom template path
            
        Returns:
            Path to the generated JSON file
        """
        # Validate input file
        self._validate_pdf_file(pdf_path)
        
        print(f"Processing PDF file: {pdf_path}")
        
        # Step 1: Read PDF file
        with open(pdf_path, 'rb') as f:
            file_bytes = f.read()
        
        # Step 2: Extract text using OCR
        print("Step 1: Extracting text using Document Intelligence...")
        extracted_text = self.ocr_processor.extract_text_from_pdf(file_bytes)
        
        # Step 3: Load template
        template = self.ai_processor.load_template(template_path)
        
        # Step 4: Process with AI to generate JSON
        print("Step 2: Processing with AI to generate structured JSON...")
        generated_json = self.ai_processor.process_text_to_json(extracted_text, template)
        
        if not generated_json:
            raise RuntimeError("AI could not process the extracted text into structured JSON")
        
        # Step 5: Save output JSON
        final_output_path = self._save_json_output(pdf_path, generated_json, output_path)
        
        # Step 6: Update XLSX file if processor is available
        xlsx_path = None
        if self.xlsx_processor:
            try:
                xlsx_path = self.xlsx_processor.update_xlsx_from_json(final_output_path)
                print(f"Updated XLSX file: {xlsx_path}")
            except Exception as e:
                print(f"Warning: Failed to update XLSX file: {e}")
        
        print(f"Successfully generated: {final_output_path}")
        return final_output_path
    
    def process_heat_number(self, heat_number: str, output_dir: Optional[str] = None, auth_token: str = None) -> str:
        """
        Process a PDF from API using HeatNumber and return the output JSON file path.
        Uses enhanced API functionality from the standalone script.
        
        Args:
            heat_number: The heat number to fetch and process
            output_dir: Optional output directory for JSON file
            auth_token: Optional authentication token for API access
            
        Returns:
            Path to the generated JSON file
        """
        if not self.api_processor:
            raise RuntimeError("API processor not available. Please check requests_pkcs12 installation and certificate configuration.")
        
        print(f"Processing HeatNumber: {heat_number}")
        
        # Fetch PDF from API with enhanced functionality
        try:
            temp_pdf_path, output_path = self.api_processor.process_heat_number_to_json(heat_number, output_dir, auth_token)
            
            # Process the temporary PDF file through the normal workflow
            result_path = self.process_pdf(temp_pdf_path, output_path)
            return result_path
            
        except Exception as e:
            raise RuntimeError(f"Failed to process HeatNumber {heat_number}: {e}")
        finally:
            # Clean up temporary PDF file if it exists
            if 'temp_pdf_path' in locals() and os.path.exists(temp_pdf_path):
                os.unlink(temp_pdf_path)
    
    def _validate_pdf_file(self, pdf_path: str):
        """Validate the input PDF file."""
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"File not found: {pdf_path}")
        
        if not pdf_path.lower().endswith('.pdf'):
            raise ValueError("Input file must be a PDF file")
    
    def _save_json_output(self, pdf_path: str, json_data: Dict[str, Any], output_path: Optional[str] = None) -> str:
        """Save the generated JSON to file."""
        if output_path:
            final_path = output_path
        else:
            # Save to Sample json folder by default
            sample_json_dir = os.path.join(os.path.dirname(__file__), "Sample json")
            os.makedirs(sample_json_dir, exist_ok=True)
            
            # Use PDF file's name but save in Sample json folder
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            final_path = os.path.join(sample_json_dir, f"{base_name}.json")
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(final_path), exist_ok=True)
        
        with open(final_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        return final_path
    
    def process_multiple_pdfs(self, pdf_paths: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple PDF files.
        
        Args:
            pdf_paths: List of PDF file paths
            output_dir: Optional output directory for all JSON files
            
        Returns:
            List of generated JSON file paths
        """
        results = []
        failed_files = []
        
        for i, pdf_path in enumerate(pdf_paths, 1):
            try:
                print(f"\nProcessing file {i}/{len(pdf_paths)}: {os.path.basename(pdf_path)}")
                
                output_path = None
                if output_dir:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    output_path = os.path.join(output_dir, f"{base_name}.json")
                
                result_path = self.process_pdf(pdf_path, output_path)
                results.append(result_path)
                
            except Exception as e:
                print(f"Error processing {pdf_path}: {e}")
                failed_files.append((pdf_path, str(e)))
        
        # Summary
        print(f"\n{'='*50}")
        print(f"Batch Processing Complete!")
        print(f"✅ Successfully processed: {len(results)} files")
        if failed_files:
            print(f"❌ Failed: {len(failed_files)} files")
            for failed_path, error in failed_files:
                print(f"   - {os.path.basename(failed_path)}: {error}")
        print(f"{'='*50}")
        
        return results


def main():
    """Main entry point for the object-oriented PDF processor."""
    print("Object-Oriented PDF Document Intelligence Processor")
    print("Converts PDF files to structured JSON using AI")
    print("=" * 60)
    
    try:
        # Initialize processor
        processor = PDFProcessor()
        
        # Interactive mode
        while True:
            print("\nSelect an option:")
            print("1. Process a single PDF file")
            print("2. Process multiple PDF files")
            print("3. Process JSON to XLSX only")
            print("4. Process PDF from API by HeatNumber")
            print("5. Exit")
            
            choice = input("\nEnter your choice (1-5): ").strip()
            
            if choice == "1":
                # Single file processing
                pdf_path = input("\nEnter the path to your PDF file: ").strip()
                
                # Remove quotes if present
                if pdf_path.startswith('"') and pdf_path.endswith('"'):
                    pdf_path = pdf_path[1:-1]
                if pdf_path.startswith("'") and pdf_path.endswith("'"):
                    pdf_path = pdf_path[1:-1]
                
                if not pdf_path:
                    print("Error: Please enter a file path.")
                    continue
                
                try:
                    # Ask for custom output path
                    use_default = input("Use default output filename? (y/n) [default: y]: ").strip().lower()
                    custom_output = None
                    if use_default == 'n':
                        custom_output = input("Enter custom output path (or press Enter for default): ").strip()
                        if not custom_output:
                            custom_output = None
                    
                    print("\nStarting processing...")
                    result_path = processor.process_pdf(pdf_path, custom_output)
                    
                    print(f"\n{'='*50}")
                    print("✅ Processing Complete!")
                    print(f"📄 Input:  {pdf_path}")
                    print(f"📋 Output: {result_path}")
                    print("="*50)
                    
                except Exception as e:
                    print(f"\nError: {e}")
            
            elif choice == "2":
                # Multiple file processing
                print("\nEnter PDF file paths (one per line, empty line to finish):")
                pdf_paths = []
                while True:
                    path = input().strip()
                    if not path:
                        break
                    # Remove quotes if present
                    if path.startswith('"') and path.endswith('"'):
                        path = path[1:-1]
                    if path.startswith("'") and path.endswith("'"):
                        path = path[1:-1]
                    pdf_paths.append(path)
                
                if not pdf_paths:
                    print("No files specified.")
                    continue
                
                # Ask for output directory
                output_dir = input("\nEnter output directory (or press Enter for same directory as input): ").strip()
                if not output_dir:
                    output_dir = None
                
                try:
                    results = processor.process_multiple_pdfs(pdf_paths, output_dir)
                    print(f"\nGenerated {len(results)} JSON files")
                except Exception as e:
                    print(f"\nError during batch processing: {e}")

            elif choice == "3":
                # JSON to XLSX conversion only
                if not processor.xlsx_processor:
                    print("Error: XLSX processing not available. Please ensure openpyxl is installed and template exists.")
                    continue
                
                json_path = input("\nEnter path to JSON file: ").strip()
                if json_path.startswith('"') and json_path.endswith('"'):
                    json_path = json_path[1:-1]
                if json_path.startswith("'") and json_path.endswith("'"):
                    json_path = json_path[1:-1]
                
                if not json_path:
                    print("Error: Please enter a JSON file path.")
                    continue
                
                try:
                    xlsx_path = processor.xlsx_processor.update_xlsx_from_json(json_path)
                    print(f"\n{'='*50}")
                    print("✅ XLSX Update Complete!")
                    print(f"📄 Input JSON:  {json_path}")
                    print(f"📊 Output XLSX: {xlsx_path}")
                    print("="*50)
                except Exception as e:
                    print(f"\nError: {e}")

            elif choice == "4":
                # Process multiple PDFs from API by HeatNumber with enhanced functionality
                if not processor.api_processor:
                    print("Error: API processing not available.")
                    if not REQUESTS_PKCS12_AVAILABLE:
                        print("Missing required package: requests_pkcs12")
                        print("Install with: pip install requests-pkcs12")
                    else:
                        print("Please check certificate configuration.")
                    continue
                
                print("\nEnter HeatNumbers to process:")
                print("Option 1: Enter multiple HeatNumbers (one per line, empty line to finish)")
                print("Option 2: Enter comma-separated HeatNumbers")
                
                input_method = input("Choose input method (1 or 2): ").strip()
                heat_numbers = []
                
                if input_method == "1":
                    print("Enter HeatNumbers (one per line, empty line to finish):")
                    while True:
                        heat_number = input().strip()
                        if not heat_number:
                            break
                        heat_numbers.append(heat_number)
                
                elif input_method == "2":
                    heat_numbers_input = input("Enter comma-separated HeatNumbers: ").strip()
                    if heat_numbers_input:
                        heat_numbers = [hn.strip() for hn in heat_numbers_input.split(",") if hn.strip()]
                
                else:
                    print("Invalid input method. Please choose 1 or 2.")
                    continue
                
                if not heat_numbers:
                    print("Error: No HeatNumbers provided.")
                    continue
                
                # Ask for authentication token (optional)
                print(f"Default token available: {DEFAULT_AUTH_TOKEN[:20]}...")
                auth_token = input("Enter auth token (or press Enter to use default): ").strip()
                if not auth_token:
                    auth_token = None
                    print("Using default authentication token")
                
                # Ask for output directory
                output_dir = input("Enter output directory (or press Enter for Sample json folder): ").strip()
                if not output_dir:
                    output_dir = None
                
                # Process multiple HeatNumbers
                try:
                    print(f"\nProcessing {len(heat_numbers)} HeatNumbers...")
                    print("=" * 60)
                    
                    successful_results = []
                    failed_results = []
                    
                    for i, heat_number in enumerate(heat_numbers, 1):
                        try:
                            print(f"\nProcessing {i}/{len(heat_numbers)}: {heat_number}")
                            print("Using enhanced API functionality...")
                            
                            result_path = processor.process_heat_number(heat_number, output_dir, auth_token)
                            successful_results.append((heat_number, result_path))
                            
                            print(f"✅ {heat_number}: Success")
                            print(f"   📋 JSON: {result_path}")
                            
                        except Exception as e:
                            failed_results.append((heat_number, str(e)))
                            print(f"❌ {heat_number}: Failed - {e}")
                    
                    # Summary
                    print(f"\n{'='*60}")
                    print("🎯 Batch API Processing Complete!")
                    print(f"✅ Successfully processed: {len(successful_results)} HeatNumbers")
                    
                    if successful_results:
                        print("\nSuccessful HeatNumbers:")
                        for heat_number, result_path in successful_results:
                            print(f"   � {heat_number} → {result_path}")
                    
                    if failed_results:
                        print(f"\n❌ Failed: {len(failed_results)} HeatNumbers")
                        for heat_number, error in failed_results:
                            print(f"   🔢 {heat_number}: {error}")
                    
                    if processor.xlsx_processor and successful_results:
                        print(f"\n📊 XLSX file updated with {len(successful_results)} records")
                    
                    print("=" * 60)
                    
                except Exception as e:
                    print(f"\nError during batch processing: {e}")
                    print("Tip: Verify your HeatNumbers and authentication token are correct.")

            elif choice == "5":
                print("\nGoodbye! 👋")
                break
            
            else:
                print("Invalid choice. Please enter 1, 2, 3, 4, or 5.")
    
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\nFatal error: {e}")
        input("\nPress Enter to exit...")
        sys.exit(1)


if __name__ == "__main__":
    main()
